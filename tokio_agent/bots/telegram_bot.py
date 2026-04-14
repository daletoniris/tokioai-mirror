"""
TokioAI Telegram Bot v2.1 — Full multimedia conversational interface.

Features:
    - Text messages -> TokioAI Agent via /chat API
    - Image/photo analysis via OpenAI Vision (Gemini fallback)
    - Voice messages -> transcription via Gemini + Whisper fallback
    - Audio files support (not just voice notes)
    - Receive files (PDF, DOCX, TXT, code) -> extract text -> TokioAI
    - Send files to user (PDF, CSV, PPTX) via send_document
    - YouTube URL detection with metadata extraction
    - Multiple images in a single message
    - Access control (ACL) by Telegram user_id
    - Owner-only admin commands (/allow, /deny, /acl)
    - Retry logic for Telegram network errors
    - Long message splitting (4000 char chunks)

Environment:
    TELEGRAM_BOT_TOKEN      - Bot token from @BotFather
    TELEGRAM_OWNER_ID       - Owner's Telegram user_id
    TELEGRAM_ALLOWED_IDS    - Comma-separated allowed user_ids
    CLI_SERVICE_URL         - TokioAI API URL (default: http://tokio-cli:8000)
    OPENAI_API_KEY          - For image/vision analysis (optional)
    OPENAI_VISION_MODEL     - Vision model (default: gpt-4o-mini)
    GEMINI_API_KEY          - For voice transcription (optional)
    TOKIO_IMAGE_MAX_BYTES   - Max image size (default: 10MB)
"""
import os
import re
import asyncio
import logging
import base64
import pathlib
import tempfile
from typing import Optional

from telegram import Update, Bot
from telegram.error import TimedOut as TelegramTimedOut, NetworkError as TelegramNetworkError
from telegram.request import HTTPXRequest
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes
import httpx

# Quick commands
from telegram_quick_cmds import (
    sitrep_command as _qc_sitrep,
    health_command as _qc_health,
    waf_command as _qc_waf,
    drone_command as _qc_drone,
    threats_command as _qc_threats,
    entity_command as _qc_entity,
    see_command as _qc_see,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
logging.getLogger("httpx").setLevel(logging.WARNING)

# ── Configuration ──
CLI_SERVICE_URL = os.getenv("CLI_SERVICE_URL", "http://tokio-cli:8000")

# Vision
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
OPENAI_VISION_MODEL = os.getenv("OPENAI_VISION_MODEL", "gpt-4o-mini").strip() or "gpt-4o-mini"
TOKIO_IMAGE_MAX_BYTES = int(os.getenv("TOKIO_IMAGE_MAX_BYTES", "10000000"))  # 10MB default

# Gemini
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()

# Uploads directory
UPLOADS_DIR = pathlib.Path(os.getenv("TOKIO_UPLOADS_DIR", "/workspace/uploads"))

# Access control
TELEGRAM_OWNER_ID = os.getenv("TELEGRAM_OWNER_ID", "").strip()
TELEGRAM_ALLOWED_IDS = os.getenv("TELEGRAM_ALLOWED_IDS", "").strip()

# YouTube URL pattern
_YOUTUBE_RE = re.compile(
    r'(?:https?://)?(?:www\.)?(?:youtube\.com/watch\?v=|youtu\.be/|youtube\.com/shorts/)([\w-]{11})',
    re.IGNORECASE,
)

user_sessions: dict = {}
allowed_user_ids: set = set()

# ── Pending file sends (agent -> user) ──
_pending_files: dict = {}  # chat_id -> list of file paths

# ── Deduplication: track recently processed update_ids ──
from collections import OrderedDict

_processed_updates: OrderedDict = OrderedDict()
_DEDUP_MAX_SIZE = 200  # keep last 200 update_ids


def _is_duplicate_update(update_id: int) -> bool:
    """Check if this update was already processed. Returns True if duplicate."""
    if update_id in _processed_updates:
        return True
    _processed_updates[update_id] = True
    # Evict oldest entries to keep memory bounded
    while len(_processed_updates) > _DEDUP_MAX_SIZE:
        _processed_updates.popitem(last=False)
    return False


# ── Helpers ──

async def _safe_send_chat_action(context: ContextTypes.DEFAULT_TYPE, chat_id: int, action: str):
    """Send chat action with retry; never let it break the flow."""
    for attempt in range(2):
        try:
            await context.bot.send_chat_action(chat_id=chat_id, action=action)
            return
        except (TelegramTimedOut, TelegramNetworkError):
            await asyncio.sleep(0.7 * (attempt + 1))
        except Exception:
            return


async def _safe_reply_text(update: Update, text: str):
    """Reply with retry logic for sporadic network errors.

    Only retries on NetworkError (connection refused, DNS, etc).
    TimedOut is NOT retried because the message may have been delivered
    despite the timeout, causing duplicate messages to the user.
    """
    if not update.message:
        return
    try:
        await update.message.reply_text(text)
    except TelegramTimedOut:
        # Timeout means Telegram might have received it — do NOT retry
        logger.warning("Telegram reply timed out — not retrying to avoid duplicate")
    except TelegramNetworkError:
        # True network failure — safe to retry once
        await asyncio.sleep(1.5)
        try:
            await update.message.reply_text(text)
        except Exception:
            logger.error("Telegram reply failed on retry")
    except Exception:
        return


async def _safe_send_document(context: ContextTypes.DEFAULT_TYPE, chat_id: int,
                               file_path: str, caption: str = ""):
    """Send a file to the user via Telegram."""
    path = pathlib.Path(file_path)
    if not path.exists():
        logger.warning(f"File not found for sending: {file_path}")
        return False
    for attempt in range(3):
        try:
            with open(path, "rb") as f:
                await context.bot.send_document(
                    chat_id=chat_id,
                    document=f,
                    filename=path.name,
                    caption=caption[:1024] if caption else None,
                )
            return True
        except (TelegramTimedOut, TelegramNetworkError):
            await asyncio.sleep(1.0 * (attempt + 1))
        except Exception as e:
            logger.error(f"Failed to send document: {e}")
            return False
    return False


def _parse_allowed_ids(raw: str) -> set:
    result = set()
    for token in (raw or "").split(","):
        token = token.strip()
        if token.isdigit():
            result.add(int(token))
    return result


def _init_access_control():
    global allowed_user_ids
    allowed_user_ids = _parse_allowed_ids(TELEGRAM_ALLOWED_IDS)
    if TELEGRAM_OWNER_ID.isdigit():
        allowed_user_ids.add(int(TELEGRAM_OWNER_ID))


def _is_owner(user_id: int) -> bool:
    return TELEGRAM_OWNER_ID.isdigit() and user_id == int(TELEGRAM_OWNER_ID)


def _is_authorized(user_id: int) -> bool:
    if not TELEGRAM_OWNER_ID and not TELEGRAM_ALLOWED_IDS:
        return True
    return user_id in allowed_user_ids


async def _guard_access(update: Update) -> bool:
    user_id = update.effective_user.id if update.effective_user else 0
    if _is_authorized(user_id):
        return True
    await _safe_reply_text(update, "⛔ No autorizado para usar este bot.")
    return False


def _get_session(user_id: int) -> str:
    if user_id not in user_sessions:
        user_sessions[user_id] = f"telegram-{user_id}"
    return user_sessions[user_id]


async def _reply_long(update: Update, text: str):
    """Reply splitting long messages into 4000-char chunks."""
    if len(text) > 4000:
        chunks = [text[i:i + 4000] for i in range(0, len(text), 4000)]
        for chunk in chunks:
            await _safe_reply_text(update, chunk)
    else:
        await _safe_reply_text(update, text)


# ── Core: Send message to TokioAI API ──

async def _send_to_tokio(message: str, session_id: str,
                         images: list = None) -> str:
    """Send a message to TokioAI /chat endpoint and return the response.

    Args:
        message: User message text.
        session_id: Session ID for continuity.
        images: Optional list of {"data": base64, "media_type": "image/..."} dicts.
    """
    timeout = httpx.Timeout(connect=10.0, read=300.0, write=30.0, pool=10.0)
    async with httpx.AsyncClient(timeout=timeout) as client:
        try:
            payload = {"message": message, "session_id": session_id}
            if images:
                payload["images"] = images
            response = await client.post(
                f"{CLI_SERVICE_URL}/chat",
                json=payload,
            )
            if response.status_code == 200:
                data = response.json()
                return data.get("response", "Sin respuesta")
            else:
                return f"Error del agente (HTTP {response.status_code}): {response.text[:300]}"
        except httpx.TimeoutException:
            return "Timeout - el agente tardo demasiado en responder."
        except httpx.RequestError as e:
            return f"Error de conexion con TokioAI: {e}"
        except Exception as e:
            return f"Error inesperado: {e}"


# ── YouTube URL handler ──

def _extract_youtube_urls(text: str) -> list:
    """Extract YouTube video IDs from text."""
    return _YOUTUBE_RE.findall(text or "")


async def _get_youtube_metadata(video_id: str) -> Optional[str]:
    """Get YouTube video metadata via yt-dlp (no download)."""
    import subprocess
    try:
        result = subprocess.run(
            ["yt-dlp", "--no-download", "--print", "%(title)s|||%(duration_string)s|||%(description).500s",
             f"https://www.youtube.com/watch?v={video_id}"],
            capture_output=True, text=True, timeout=15,
        )
        if result.returncode == 0 and result.stdout.strip():
            parts = result.stdout.strip().split("|||")
            title = parts[0] if len(parts) > 0 else "?"
            duration = parts[1] if len(parts) > 1 else "?"
            desc = parts[2] if len(parts) > 2 else ""
            return f"YouTube: {title} ({duration})\nDescripcion: {desc[:300]}"
    except Exception as e:
        logger.debug(f"yt-dlp failed: {e}")
    return None


# ── Handlers ──

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _guard_access(update):
        return
    user = update.effective_user
    await _safe_reply_text(
        update,
        f"Hola {user.first_name}!\n\n"
        "Soy TokioAI v2.1 - Agente Autonomo.\n\n"
        "Puedo ayudarte con:\n"
        "- Seguridad WAF y gestion de tenants\n"
        "- Docker y contenedores\n"
        "- Infraestructura GCP\n"
        "- IoT y domotica\n"
        "- Calendario y productividad\n"
        "- Automatizacion de tareas\n"
        "- Analisis de imagenes\n"
        "- Notas de voz y audio\n"
        "- Documentos (PDF, DOCX, CSV)\n"
        "- URLs de YouTube\n"
        "- Generacion de reportes PDF/Slides\n\n"
        "Simplemente escribe tu comando o pregunta."
    )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle text messages -> TokioAI. Detects YouTube URLs."""
    if not await _guard_access(update):
        return
    # Deduplicate: skip if we already processed this update
    if _is_duplicate_update(update.update_id):
        logger.warning(f"⚠️ Duplicate update_id {update.update_id} — skipping")
        return
    user_id = update.effective_user.id
    message_text = update.message.text
    logger.info(f"User {user_id}: {message_text}")

    await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    session_id = _get_session(user_id)

    # Check for YouTube URLs
    yt_ids = _extract_youtube_urls(message_text)
    yt_context = ""
    if yt_ids:
        for vid in yt_ids[:3]:  # Max 3 videos
            meta = await _get_youtube_metadata(vid)
            if meta:
                yt_context += f"\n\n[Video YouTube detectado]\n{meta}\nURL: https://www.youtube.com/watch?v={vid}\n"

    prompt = message_text
    if yt_context:
        prompt = f"{message_text}\n{yt_context}"

    # Keep sending typing while waiting
    async def _keep_typing():
        while True:
            await asyncio.sleep(5)
            await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    typing_task = asyncio.create_task(_keep_typing())
    try:
        result = await _send_to_tokio(prompt, session_id)
    finally:
        typing_task.cancel()

    # Check if agent generated files to send
    await _send_pending_files(context, update.effective_chat.id, result)

    await _reply_long(update, result)


# ── Image handling ──

def _detect_image_mime(image_bytes: bytes) -> str:
    if image_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if image_bytes.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if image_bytes[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if image_bytes.startswith(b"RIFF") and b"WEBP" in image_bytes[:16]:
        return "image/webp"
    return "application/octet-stream"


async def _download_telegram_file(context: ContextTypes.DEFAULT_TYPE, file_id: str) -> bytes:
    tg_file = await context.bot.get_file(file_id)
    for method_name in ("download_as_bytearray", "download_as_bytes"):
        method = getattr(tg_file, method_name, None)
        if method:
            data = await method()
            return bytes(data)
    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        tmp_path = tmp.name
    try:
        await tg_file.download_to_drive(tmp_path)
        return pathlib.Path(tmp_path).read_bytes()
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


async def _download_telegram_file_to_disk(context: ContextTypes.DEFAULT_TYPE,
                                            file_id: str, dest_path: str) -> bool:
    """Download a Telegram file directly to disk."""
    tg_file = await context.bot.get_file(file_id)
    try:
        await tg_file.download_to_drive(dest_path)
        return True
    except Exception as e:
        logger.error(f"Failed to download file to {dest_path}: {e}")
        return False


async def _openai_vision_to_text(image_bytes: bytes, caption: str = "") -> str:
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY no configurado")

    mime = _detect_image_mime(image_bytes)
    if mime == "application/octet-stream":
        mime = "image/png"

    if len(image_bytes) > TOKIO_IMAGE_MAX_BYTES:
        raise RuntimeError(f"Imagen muy grande ({len(image_bytes)} bytes). Limite: {TOKIO_IMAGE_MAX_BYTES}.")

    data_url = f"data:{mime};base64,{base64.b64encode(image_bytes).decode('ascii')}"
    user_text = (
        "Analiza la imagen adjunta y responde en Espanol.\n"
        "1) Describe lo que ves (resumen).\n"
        "2) Extrae TODO el texto visible (OCR) tal cual.\n"
        "3) Si hay informacion tecnica (logs, errores, pantallas), explica lo importante.\n"
        "4) Si hay indicadores de seguridad (IPs, dominios, comandos), listalos.\n"
    )
    if caption:
        user_text += f"\nContexto del usuario: {caption}\n"

    payload = {
        "model": OPENAI_VISION_MODEL,
        "messages": [
            {"role": "system", "content": "Eres un analista experto. Se conciso y preciso."},
            {"role": "user", "content": [
                {"type": "text", "text": user_text},
                {"type": "image_url", "image_url": {"url": data_url}},
            ]},
        ],
        "max_tokens": 900,
        "temperature": 0.2,
    }

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
            json=payload,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"OpenAI vision error: {resp.status_code} {resp.text[:200]}")
        data = resp.json()
        return (data["choices"][0]["message"]["content"] or "").strip()


async def _gemini_vision_to_text(image_bytes: bytes, caption: str = "") -> str:
    """Fallback: analyze image via Gemini REST API (no SDK dependency)."""
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY no configurado")

    mime = _detect_image_mime(image_bytes)
    if mime == "application/octet-stream":
        mime = "image/png"

    b64_data = base64.b64encode(image_bytes).decode("ascii")

    prompt = "Analiza esta imagen. Describe lo que ves, extrae texto visible (OCR), y resalta info tecnica."
    if caption:
        prompt += f"\nContexto: {caption}"

    payload = {
        "contents": [{
            "parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": mime, "data": b64_data}},
            ]
        }],
        "generationConfig": {"temperature": 0.2, "maxOutputTokens": 1024},
    }

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"

    async with httpx.AsyncClient(timeout=60.0) as client:
        resp = await client.post(url, json=payload)
        if resp.status_code != 200:
            raise RuntimeError(f"Gemini vision error: {resp.status_code} {resp.text[:200]}")
        data = resp.json()
        candidates = data.get("candidates", [])
        if not candidates:
            raise RuntimeError("Gemini: no candidates in response")
        parts = candidates[0].get("content", {}).get("parts", [])
        text = "".join(p.get("text", "") for p in parts).strip()
        if not text:
            raise RuntimeError("Gemini: empty response text")
        return text


async def handle_image(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle photo messages — sends image directly to the agent's LLM (Claude Vision)."""
    if not await _guard_access(update):
        return
    if not update.message or not update.effective_user:
        return

    user_id = update.effective_user.id
    caption = (update.message.caption or "").strip()

    # Collect image file IDs
    image_file_ids = []
    if update.message.photo:
        candidates = [p for p in update.message.photo if getattr(p, "file_id", None)]
        if candidates:
            chosen = candidates[-1]
            for p in reversed(candidates):
                fs = getattr(p, "file_size", None)
                if fs and fs <= TOKIO_IMAGE_MAX_BYTES:
                    chosen = p
                    break
            image_file_ids.append(chosen.file_id)

    if not image_file_ids:
        return

    await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    # Download images and encode as base64
    images_payload = []
    for file_id in image_file_ids:
        try:
            image_bytes = await _download_telegram_file(context, file_id)
            mime = _detect_image_mime(image_bytes)
            if mime == "application/octet-stream":
                mime = "image/jpeg"
            b64 = base64.b64encode(image_bytes).decode("ascii")
            images_payload.append({"data": b64, "media_type": mime})
        except Exception as e:
            logger.error(f"Failed to download image: {e}")

    if not images_payload:
        await _safe_reply_text(update, "No pude descargar la imagen. Reintenta.")
        return

    # Build prompt
    prompt = caption if caption else "Analiza esta imagen. Describe lo que ves, extrae texto visible (OCR), y resalta informacion tecnica o de seguridad."

    session_id = _get_session(user_id)

    # Keep sending typing while waiting
    async def _keep_typing():
        while True:
            await asyncio.sleep(5)
            await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    typing_task = asyncio.create_task(_keep_typing())
    try:
        result = await _send_to_tokio(prompt, session_id, images=images_payload)
    finally:
        typing_task.cancel()

    # Check for files to send
    await _send_pending_files(context, update.effective_chat.id, result)

    await _reply_long(update, result)


# ── Video handling (extract frames → Claude Vision) ──

TOKIO_VIDEO_MAX_BYTES = int(os.getenv("TOKIO_VIDEO_MAX_BYTES", "20000000"))  # 20MB default
TOKIO_VIDEO_MAX_FRAMES = int(os.getenv("TOKIO_VIDEO_MAX_FRAMES", "4"))  # max frames to extract


def _extract_video_frames_sync(video_path: str, max_frames: int = 4) -> list:
    """Extract key frames from a video using ffmpeg (synchronous).

    Returns list of (frame_bytes, mime_type) tuples.
    """
    import subprocess as sp
    frames = []

    # Get video duration
    try:
        probe = sp.run(
            ["ffprobe", "-v", "quiet", "-show_entries", "format=duration",
             "-of", "csv=p=0", video_path],
            capture_output=True, text=True, timeout=10,
        )
        duration = float(probe.stdout.strip() or "5")
    except Exception:
        duration = 5.0

    # Calculate timestamps to extract (evenly spaced)
    if duration <= 0:
        duration = 5.0
    n = min(max_frames, max(1, int(duration // 2)))  # 1 frame per 2 seconds, max 4
    timestamps = [duration * (i + 1) / (n + 1) for i in range(n)]

    for i, ts in enumerate(timestamps):
        out_path = f"{video_path}_frame{i}.jpg"
        try:
            sp.run(
                ["ffmpeg", "-y", "-ss", str(ts), "-i", video_path,
                 "-frames:v", "1", "-q:v", "2", out_path],
                capture_output=True, timeout=15,
            )
            frame_data = pathlib.Path(out_path).read_bytes()
            if len(frame_data) > 100:  # sanity check
                frames.append((frame_data, "image/jpeg"))
        except Exception as e:
            logger.debug(f"Frame extraction failed at {ts}s: {e}")
        finally:
            try:
                os.unlink(out_path)
            except Exception:
                pass

    return frames


async def handle_video(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle video/video_note messages — extract frames and analyze with Claude Vision."""
    if not await _guard_access(update):
        return
    if not update.message or not update.effective_user:
        return

    user_id = update.effective_user.id
    caption = (update.message.caption or "").strip()

    # Get video file_id (supports video, video_note, animation)
    video = update.message.video or update.message.video_note or update.message.animation
    if not video:
        return

    file_size = getattr(video, "file_size", 0) or 0
    if file_size > TOKIO_VIDEO_MAX_BYTES:
        await _safe_reply_text(
            update,
            f"Video demasiado grande ({file_size // 1_000_000}MB). "
            f"Maximo: {TOKIO_VIDEO_MAX_BYTES // 1_000_000}MB"
        )
        return

    await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    # Download video to temp file
    try:
        video_bytes = await _download_telegram_file(context, video.file_id)
    except Exception as e:
        await _safe_reply_text(update, f"No pude descargar el video: {e}")
        return

    # Save to temp and extract frames
    with tempfile.NamedTemporaryFile(suffix=".mp4", delete=False) as tmp:
        tmp.write(video_bytes)
        tmp_path = tmp.name

    try:
        loop = asyncio.get_event_loop()
        frames = await loop.run_in_executor(
            None, _extract_video_frames_sync, tmp_path, TOKIO_VIDEO_MAX_FRAMES
        )
    except Exception as e:
        logger.error(f"Video frame extraction failed: {e}")
        frames = []

    try:
        os.unlink(tmp_path)
    except Exception:
        pass

    if not frames:
        await _safe_reply_text(update, "No pude extraer frames del video. Necesito ffmpeg instalado.")
        return

    # Build images payload
    images_payload = []
    for frame_bytes, mime in frames:
        b64 = base64.b64encode(frame_bytes).decode("ascii")
        images_payload.append({"data": b64, "media_type": mime})

    # Build prompt
    duration_info = f" ({len(frames)} frames extraidos)"
    if caption:
        prompt = f"{caption}\n\n[Video recibido{duration_info}. Estas son capturas del video.]"
    else:
        prompt = (
            f"Analiza este video{duration_info}. "
            "Describe lo que ves en cada frame, detecta movimiento, personas, objetos, "
            "texto visible y cualquier informacion relevante de seguridad. "
            "Las imagenes son frames extraidos del video en orden cronologico."
        )

    session_id = _get_session(user_id)

    async def _keep_typing():
        while True:
            await asyncio.sleep(5)
            await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    typing_task = asyncio.create_task(_keep_typing())
    try:
        result = await _send_to_tokio(prompt, session_id, images=images_payload)
    finally:
        typing_task.cancel()

    await _send_pending_files(context, update.effective_chat.id, result)
    await _reply_long(update, result)


# ── Document handling (receive files from user) ──

_TEXT_EXTENSIONS = {
    ".txt", ".md", ".log", ".json", ".yaml", ".yml", ".xml", ".csv", ".tsv",
    ".py", ".js", ".ts", ".go", ".rs", ".java", ".c", ".cpp", ".h", ".sh",
    ".bash", ".zsh", ".sql", ".html", ".css", ".conf", ".ini", ".toml",
    ".env", ".dockerfile", ".makefile",
}


def _extract_pdf_text(path: str) -> str:
    """Extract text from PDF using PyPDF2 or pdfplumber."""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(path)
        text_parts = []
        for page in reader.pages[:50]:  # Limit to 50 pages
            t = page.extract_text()
            if t:
                text_parts.append(t)
        return "\n\n".join(text_parts)
    except ImportError:
        pass
    try:
        import pdfplumber
        with pdfplumber.open(path) as pdf:
            text_parts = []
            for page in pdf.pages[:50]:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
            return "\n\n".join(text_parts)
    except ImportError:
        pass
    return "[No se pudo extraer texto del PDF - falta PyPDF2 o pdfplumber]"


def _extract_docx_text(path: str) -> str:
    """Extract text from DOCX."""
    try:
        from docx import Document
        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except ImportError:
        return "[No se pudo extraer texto del DOCX - falta python-docx]"
    except Exception as e:
        return f"[Error extrayendo DOCX: {e}]"


def _extract_xlsx_text(path: str) -> str:
    """Extract text from XLSX."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        parts = []
        for ws in wb.worksheets[:5]:  # Max 5 sheets
            rows = []
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append(" | ".join(cells))
            parts.append(f"[Hoja: {ws.title}]\n" + "\n".join(rows))
        return "\n\n".join(parts)
    except ImportError:
        return "[No se pudo extraer texto del XLSX - falta openpyxl]"
    except Exception as e:
        return f"[Error extrayendo XLSX: {e}]"


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle document uploads - extract text and send to TokioAI."""
    if not await _guard_access(update):
        return
    if not update.message or not update.message.document or not update.effective_user:
        return

    doc = update.message.document
    mime = (doc.mime_type or "").lower()
    file_name = doc.file_name or "unknown"
    ext = pathlib.Path(file_name).suffix.lower()
    caption = (update.message.caption or "").strip()
    user_id = update.effective_user.id

    # Check if it's an image sent as document
    if mime.startswith("image/"):
        await handle_image(update, context)
        return

    await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="upload_document")

    # Create uploads directory
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    dest_path = UPLOADS_DIR / file_name

    try:
        ok = await _download_telegram_file_to_disk(context, doc.file_id, str(dest_path))
        if not ok:
            await _safe_reply_text(update, "No pude descargar el archivo. Reintenta.")
            return
    except Exception as e:
        logger.error(f"Document download failed: {e}")
        await _safe_reply_text(update, "Error descargando archivo.")
        return

    await _safe_reply_text(update, f"Archivo recibido: {file_name}. Procesando...")

    # Extract text based on file type
    extracted = ""
    if ext == ".pdf" or mime == "application/pdf":
        extracted = await asyncio.to_thread(_extract_pdf_text, str(dest_path))
    elif ext in (".docx",) or mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        extracted = await asyncio.to_thread(_extract_docx_text, str(dest_path))
    elif ext in (".xlsx",) or mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        extracted = await asyncio.to_thread(_extract_xlsx_text, str(dest_path))
    elif ext in _TEXT_EXTENSIONS or mime.startswith("text/"):
        try:
            extracted = dest_path.read_text(errors="replace")[:50000]
        except Exception as e:
            extracted = f"[Error leyendo archivo: {e}]"
    else:
        extracted = f"[Archivo binario: {file_name} ({mime}), guardado en {dest_path}]"

    # Truncate very large text
    if len(extracted) > 30000:
        extracted = extracted[:30000] + "\n\n[... truncado ...]"

    command = (
        f"El usuario envio un archivo: {file_name}\n"
        f"Tipo: {mime}\n"
        f"Guardado en: {dest_path}\n\n"
        f"CONTENIDO EXTRAIDO:\n{extracted}\n\n"
    )
    if caption:
        command += f"INSTRUCCION DEL USUARIO: {caption}\n"
    else:
        command += "INSTRUCCION: Analiza el contenido del archivo y responde al usuario.\n"

    session_id = _get_session(user_id)
    result = await _send_to_tokio(command, session_id)
    await _send_pending_files(context, update.effective_chat.id, result)
    await _reply_long(update, result)


# ── Voice / Audio handling ──

async def _transcribe_with_gemini(audio_path: str) -> Optional[str]:
    """Transcribe audio using Gemini REST API."""
    if not GEMINI_API_KEY:
        return None
    try:
        audio_bytes = pathlib.Path(audio_path).read_bytes()
        b64_data = base64.b64encode(audio_bytes).decode("ascii")

        # Detect audio MIME type
        ext = pathlib.Path(audio_path).suffix.lower()
        mime_map = {".mp3": "audio/mp3", ".wav": "audio/wav", ".ogg": "audio/ogg",
                    ".m4a": "audio/mp4", ".flac": "audio/flac"}
        mime = mime_map.get(ext, "audio/mp3")

        payload = {
            "contents": [{
                "parts": [
                    {"text": "Transcribe this audio to text in Spanish. Only return the transcription."},
                    {"inline_data": {"mime_type": mime, "data": b64_data}},
                ]
            }],
            "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048},
        }

        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}"

        async with httpx.AsyncClient(timeout=60.0) as client:
            resp = await client.post(url, json=payload)
            if resp.status_code != 200:
                logger.debug(f"Gemini transcription HTTP {resp.status_code}: {resp.text[:200]}")
                return None
            data = resp.json()
            candidates = data.get("candidates", [])
            if not candidates:
                return None
            parts = candidates[0].get("content", {}).get("parts", [])
            text = "".join(p.get("text", "") for p in parts).strip()
            return text if text else None
    except Exception as e:
        logger.debug(f"Gemini transcription failed: {e}")
        return None


async def _transcribe_with_whisper(audio_path: str) -> Optional[str]:
    """Transcribe audio using OpenAI Whisper as fallback."""
    if not OPENAI_API_KEY:
        return None
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            with open(audio_path, "rb") as f:
                resp = await client.post(
                    "https://api.openai.com/v1/audio/transcriptions",
                    headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
                    files={"file": (pathlib.Path(audio_path).name, f)},
                    data={"model": "whisper-1", "language": "es"},
                )
            if resp.status_code == 200:
                data = resp.json()
                return data.get("text", "").strip() or None
            logger.debug(f"Whisper API error: {resp.status_code}")
    except Exception as e:
        logger.debug(f"Whisper transcription failed: {e}")
    return None


async def _convert_audio(input_path: str) -> Optional[str]:
    """Convert audio to mp3 or wav using ffmpeg."""
    import subprocess
    mp3_path = input_path.rsplit(".", 1)[0] + ".mp3"
    result = subprocess.run(
        ["ffmpeg", "-y", "-i", input_path, "-ar", "16000", "-ac", "1", "-f", "mp3", mp3_path],
        capture_output=True, check=False,
    )
    if result.returncode == 0 and os.path.exists(mp3_path):
        return mp3_path
    # Try WAV
    wav_path = input_path.rsplit(".", 1)[0] + ".wav"
    result = subprocess.run(
        ["ffmpeg", "-y", "-i", input_path, "-ar", "16000", "-ac", "1", wav_path],
        capture_output=True, check=False,
    )
    if result.returncode == 0 and os.path.exists(wav_path):
        return wav_path
    return None


async def voice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle voice messages - transcribe via Gemini + Whisper fallback."""
    if not await _guard_access(update):
        return
    user_id = update.effective_user.id
    voice = update.message.voice

    await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    if not GEMINI_API_KEY and not OPENAI_API_KEY:
        await _safe_reply_text(
            update,
            "Nota de voz recibida.\n\n"
            "Para procesar audios, configura GEMINI_API_KEY o OPENAI_API_KEY.\n"
            "O envia el mensaje como texto."
        )
        return

    tmp_path = None
    audio_path = None
    try:
        file = await context.bot.get_file(voice.file_id)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".ogg") as tmp_file:
            await file.download_to_drive(tmp_file.name)
            tmp_path = tmp_file.name

        audio_path = await asyncio.to_thread(_convert_audio_sync, tmp_path)
        if not audio_path:
            await _safe_reply_text(update, "No pude convertir el audio. Envia como texto.")
            return

        # Try Gemini first, then Whisper
        transcription = await _transcribe_with_gemini(audio_path)
        if not transcription:
            transcription = await _transcribe_with_whisper(audio_path)

        if not transcription:
            await _safe_reply_text(update, "No se pudo transcribir el audio. Envia como texto.")
            return

        await _safe_reply_text(update, f"Transcrito: {transcription}\n\nProcesando...")

        session_id = _get_session(user_id)
        result = await _send_to_tokio(transcription, session_id)
        await _send_pending_files(context, update.effective_chat.id, result)
        await _reply_long(update, result)

    except asyncio.TimeoutError:
        await _safe_reply_text(update, "Timeout transcribiendo audio. Envia como texto.")
    except Exception as e:
        logger.error(f"Voice handler error: {e}", exc_info=True)
        await _safe_reply_text(update, f"Error procesando audio: {str(e)[:200]}")
    finally:
        for p in [tmp_path, audio_path]:
            if p:
                try:
                    os.unlink(p)
                except Exception:
                    pass


def _convert_audio_sync(input_path: str) -> Optional[str]:
    """Synchronous audio conversion helper."""
    import subprocess
    mp3_path = input_path.replace(".ogg", ".mp3")
    result = subprocess.run(
        ["ffmpeg", "-y", "-i", input_path, "-ar", "16000", "-ac", "1", "-f", "mp3", mp3_path],
        capture_output=True, check=False,
    )
    if result.returncode == 0 and os.path.exists(mp3_path):
        return mp3_path
    wav_path = input_path.replace(".ogg", ".wav")
    result = subprocess.run(
        ["ffmpeg", "-y", "-i", input_path, "-ar", "16000", "-ac", "1", wav_path],
        capture_output=True, check=False,
    )
    if result.returncode == 0 and os.path.exists(wav_path):
        return wav_path
    return None


async def audio_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handle audio files (MP3, WAV, etc. — not voice notes)."""
    if not await _guard_access(update):
        return
    user_id = update.effective_user.id
    audio = update.message.audio

    await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

    if not GEMINI_API_KEY and not OPENAI_API_KEY:
        await _safe_reply_text(
            update,
            "Audio recibido. Configura GEMINI_API_KEY o OPENAI_API_KEY para transcribir."
        )
        return

    tmp_path = None
    try:
        suffix = pathlib.Path(audio.file_name or "audio.mp3").suffix or ".mp3"
        file = await context.bot.get_file(audio.file_id)
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp_file:
            await file.download_to_drive(tmp_file.name)
            tmp_path = tmp_file.name

        # Transcribe: Gemini first, Whisper fallback
        transcription = await _transcribe_with_gemini(tmp_path)
        if not transcription:
            transcription = await _transcribe_with_whisper(tmp_path)

        if not transcription:
            await _safe_reply_text(update, "No pude transcribir el audio.")
            return

        title = audio.title or audio.file_name or "Audio"
        await _safe_reply_text(update, f"Audio '{title}' transcrito:\n{transcription}\n\nProcesando...")

        session_id = _get_session(user_id)
        result = await _send_to_tokio(transcription, session_id)
        await _reply_long(update, result)

    except Exception as e:
        logger.error(f"Audio handler error: {e}", exc_info=True)
        await _safe_reply_text(update, f"Error procesando audio: {str(e)[:200]}")
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


# ── File sending (agent -> user) ──

# Regex to find file paths in agent responses
_FILE_PATH_RE = re.compile(r'(/workspace/[^\s\n\r\`\'"<>|*?]+\.(?:pdf|csv|pptx|xlsx|docx|txt|json|png|jpg|zip|tar\.gz))', re.IGNORECASE)


async def _send_pending_files(context: ContextTypes.DEFAULT_TYPE, chat_id: int,
                                response_text: str = ""):
    """Send files referenced in the agent response + any in the output directory."""
    sent_paths = set()

    # 1. Extract file paths from the agent's response text
    if response_text:
        matches = _FILE_PATH_RE.findall(response_text)
        for fpath in matches:
            fpath = fpath.rstrip(".,;:)")
            p = pathlib.Path(fpath)
            if p.exists() and p.is_file() and p.stat().st_size > 0 and str(p) not in sent_paths:
                sent = await _safe_send_document(context, chat_id, str(p),
                                                  caption=f"{p.name}")
                if sent:
                    sent_paths.add(str(p))
                    logger.info(f"Sent file to user: {p.name}")

    # 2. Also check the output directory for any generated files
    output_dir = pathlib.Path("/workspace/output")
    if output_dir.exists():
        for f in sorted(output_dir.iterdir()):
            if f.is_file() and f.stat().st_size > 0 and str(f) not in sent_paths:
                sent = await _safe_send_document(context, chat_id, str(f),
                                                  caption=f"{f.name}")
                if sent:
                    sent_paths.add(str(f))
                    try:
                        f.unlink()
                    except Exception:
                        pass


# ── Admin commands ──

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _guard_access(update):
        return
    await _safe_reply_text(
        update,
        "AYUDA - TokioAI Bot v2.1\n\n"
        "Comandos:\n"
        "/start - Iniciar bot\n"
        "/help - Esta ayuda\n"
        "/status - Estado del servicio\n"
        "/tools - Herramientas disponibles\n\n"
        "Admin (solo owner):\n"
        "/myid - Ver tu user_id/chat_id\n"
        "/allow <id> - Autorizar usuario\n"
        "/deny <id> - Revocar usuario\n"
        "/acl - Ver ACL actual\n\n"
        "Multimedia soportado:\n"
        "- Fotos/imagenes (analisis con IA)\n"
        "- Notas de voz y archivos de audio\n"
        "- Documentos (PDF, DOCX, XLSX, TXT, codigo)\n"
        "- URLs de YouTube (metadata automatica)\n\n"
        "Ejemplos:\n"
        "- 'muestrame los contenedores docker'\n"
        "- 'agrega el sitio ejemplo.com'\n"
        "- 'enciende las luces de la sala'\n"
        "- 'hazme un reporte PDF del WAF'\n"
        "- Envia una foto, documento o nota de voz"
    )


async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Combined status: health + engine stats + features."""
    if not await _guard_access(update):
        return
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            # Fetch health and stats in parallel
            health_resp, stats_resp = await asyncio.gather(
                client.get(f"{CLI_SERVICE_URL}/health"),
                client.get(f"{CLI_SERVICE_URL}/stats"),
                return_exceptions=True,
            )

            lines = ["📊 TokioAI Status\n"]

            # Health info
            if isinstance(health_resp, Exception):
                lines.append("❌ Agent: no responde")
            elif health_resp.status_code == 200:
                h = health_resp.json()
                lines.append(f"✅ Agent: {h.get('status', '?')}")
                lines.append(f"🤖 LLM: {h.get('llm', '?')}")
                lines.append(f"🔧 Tools: {h.get('tools', '?')}")
            else:
                lines.append(f"⚠️ Agent: HTTP {health_resp.status_code}")

            # Stats info
            if not isinstance(stats_resp, Exception) and stats_resp.status_code == 200:
                s = stats_resp.json()
                tokens = s.get("total_tokens", 0)
                if tokens > 0:
                    lines.append(f"📝 Tokens: {tokens:,}")
                tools_exec = s.get("tools_executed", 0)
                if tools_exec > 0:
                    lines.append(f"⚡ Tools ejecutadas: {tools_exec}")
                compactions = s.get("compactions", 0)
                if compactions > 0:
                    lines.append(f"🗜 Compactaciones: {compactions}")
                mem = s.get("auto_memory", {})
                mem_saved = mem.get("total_memories_saved", 0)
                if mem_saved > 0:
                    lines.append(f"🧠 Memorias: {mem_saved}")
                sub = s.get("subagents", {})
                if sub.get("total_spawned", 0) > 0:
                    lines.append(f"👷 Workers: {sub['total_spawned']} total, {sub.get('running', 0)} activos")

            # Engine features
            lines.append("\n🚀 Features activas:")
            lines.append("  • Auto-compact (context management)")
            lines.append("  • Auto-memory (persistent learning)")
            lines.append("  • Skills (/commands)")
            lines.append("  • Subagents (parallel workers)")
            lines.append("  • Structured file editing")

            await _safe_reply_text(update, "\n".join(lines))

    except Exception as e:
        await _safe_reply_text(update, f"Error: {str(e)[:200]}")


async def tools_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _guard_access(update):
        return
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            resp = await client.get(f"{CLI_SERVICE_URL}/tools")
            if resp.status_code == 200:
                data = resp.json()
                tools = data.get("tools", [])
                text = f"Herramientas: {len(tools)}\n\n"
                categories: dict = {}
                for tool in tools:
                    cat = tool.get("category", "General")
                    categories.setdefault(cat, []).append(tool["name"])
                for cat, names in sorted(categories.items()):
                    text += f"{cat}:\n"
                    for name in names:
                        text += f"  - {name}\n"
                    text += "\n"
                await _safe_reply_text(update, text)
            else:
                await _safe_reply_text(update, "No se pudieron obtener las herramientas")
    except Exception as e:
        await _safe_reply_text(update, f"Error: {str(e)[:200]}")


async def myid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await _guard_access(update):
        return
    user_id = update.effective_user.id if update.effective_user else None
    chat_id = update.effective_chat.id if update.effective_chat else None
    await _safe_reply_text(update, f"user_id={user_id}\nchat_id={chat_id}")


async def allow_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user or not _is_owner(update.effective_user.id):
        await _safe_reply_text(update, "Solo el owner puede ejecutar /allow")
        return
    if not context.args or not context.args[0].isdigit():
        await _safe_reply_text(update, "Uso: /allow <telegram_user_id>")
        return
    allowed_user_ids.add(int(context.args[0]))
    await _safe_reply_text(update, f"Usuario {context.args[0]} autorizado")


async def deny_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user or not _is_owner(update.effective_user.id):
        await _safe_reply_text(update, "Solo el owner puede ejecutar /deny")
        return
    if not context.args or not context.args[0].isdigit():
        await _safe_reply_text(update, "Uso: /deny <telegram_user_id>")
        return
    denied_id = int(context.args[0])
    if TELEGRAM_OWNER_ID.isdigit() and denied_id == int(TELEGRAM_OWNER_ID):
        await _safe_reply_text(update, "No puedes remover al owner")
        return
    allowed_user_ids.discard(denied_id)
    await _safe_reply_text(update, f"Usuario {denied_id} removido")


async def acl_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.effective_user or not _is_owner(update.effective_user.id):
        await _safe_reply_text(update, "Solo el owner puede ejecutar /acl")
        return
    owner_text = TELEGRAM_OWNER_ID if TELEGRAM_OWNER_ID else "(no configurado)"
    await _safe_reply_text(
        update,
        f"ACL\nOwner: {owner_text}\nPermitidos: {sorted(list(allowed_user_ids))}"
    )


_conflict_count = 0

async def _error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    global _conflict_count
    from telegram.error import Conflict as TelegramConflict
    error = context.error
    if isinstance(error, TelegramConflict):
        _conflict_count += 1
        if _conflict_count <= 5:
            logger.warning(
                f"Conflict error #{_conflict_count} - another polling instance. "
                f"Waiting 10s..."
            )
            await asyncio.sleep(10)
        else:
            logger.error(
                f"Conflict error #{_conflict_count} - persistent conflict. "
                f"Another bot instance is likely running with the same token. "
                f"Suppressing further conflict logs."
            )
            await asyncio.sleep(30)
        return
    # Reset conflict counter on any other error type (means polling is working)
    _conflict_count = 0
    if isinstance(error, (TelegramTimedOut, TelegramNetworkError)):
        logger.warning(f"Network error: {error}")
        return
    logger.error(f"Unhandled error: {error}", exc_info=context.error)


# ── Main ──

def main():
    import time

    token = os.getenv("TELEGRAM_BOT_TOKEN", "").strip()
    if not token:
        logger.error("TELEGRAM_BOT_TOKEN not set. Create bot with @BotFather.")
        return

    _init_access_control()
    logger.info(f"ACL: {len(allowed_user_ids)} usuarios autorizados")

    # No manual Telegram API calls here — run_polling(drop_pending_updates=True)
    # handles deleteWebhook and getUpdates internally.
    # Manual calls were causing 409 Conflict with the library's own polling.

    logger.info("TokioAI Telegram Bot v2.1 starting...")
    logger.info(f"CLI Service: {CLI_SERVICE_URL}")

    # Ensure uploads directory exists
    UPLOADS_DIR.mkdir(parents=True, exist_ok=True)

    request = HTTPXRequest(
        connect_timeout=20.0,
        read_timeout=45.0,
        write_timeout=30.0,
        pool_timeout=20.0,
    )
    application = Application.builder().token(token).request(request).build()

    application.add_error_handler(_error_handler)

    # Command handlers
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("help", help_command))
    application.add_handler(CommandHandler("status", status_command))
    application.add_handler(CommandHandler("tools", tools_command))
    application.add_handler(CommandHandler("myid", myid_command))
    application.add_handler(CommandHandler("allow", allow_command))
    application.add_handler(CommandHandler("deny", deny_command))
    application.add_handler(CommandHandler("acl", acl_command))

    # Quick command handlers (instant, no LLM)
    async def _wrap_sitrep(u, c):
        if not await _guard_access(u): return
        await _qc_sitrep(u, c)
    async def _wrap_health(u, c):
        if not await _guard_access(u): return
        await _qc_health(u, c)
    async def _wrap_waf(u, c):
        if not await _guard_access(u): return
        await _qc_waf(u, c)
    async def _wrap_drone(u, c):
        if not await _guard_access(u): return
        await _qc_drone(u, c)
    async def _wrap_threats(u, c):
        if not await _guard_access(u): return
        await _qc_threats(u, c)
    async def _wrap_entity(u, c):
        if not await _guard_access(u): return
        await _qc_entity(u, c)
    async def _wrap_see(u, c):
        if not await _guard_access(u): return
        await _qc_see(u, c)
    application.add_handler(CommandHandler("sitrep", _wrap_sitrep))
    application.add_handler(CommandHandler("health", _wrap_health))
    application.add_handler(CommandHandler("waf", _wrap_waf))
    application.add_handler(CommandHandler("drone", _wrap_drone))
    application.add_handler(CommandHandler("threats", _wrap_threats))
    application.add_handler(CommandHandler("entity", _wrap_entity))
    application.add_handler(CommandHandler("see", _wrap_see))

    # Catch-all for unrecognized /commands — pass them to the engine as skills
    async def unknown_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Route unrecognized /commands to the engine's skill system."""
        if not await _guard_access(update):
            return
        if _is_duplicate_update(update.update_id):
            return
        user_id = update.effective_user.id
        session_id = _get_session(user_id)
        message_text = update.message.text
        logger.info(f"User {user_id} skill: {message_text}")

        await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

        async def _keep_typing():
            while True:
                await asyncio.sleep(5)
                await _safe_send_chat_action(context, chat_id=update.effective_chat.id, action="typing")

        typing_task = asyncio.create_task(_keep_typing())
        try:
            result = await _send_to_tokio(message_text, session_id)
        finally:
            typing_task.cancel()

        await _reply_long(update, result)

    application.add_handler(MessageHandler(filters.COMMAND, unknown_command))

    # Content handlers — order matters!
    application.add_handler(MessageHandler(filters.PHOTO, handle_image))
    application.add_handler(MessageHandler(filters.VIDEO | filters.VIDEO_NOTE | filters.ANIMATION, handle_video))
    application.add_handler(MessageHandler(filters.VOICE, voice_handler))
    application.add_handler(MessageHandler(filters.AUDIO, audio_handler))
    application.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    application.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=True,
        poll_interval=3.0,
        timeout=30,
    )


if __name__ == "__main__":
    main()
